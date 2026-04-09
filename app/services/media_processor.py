"""
MediaProcessor — interpreta archivos adjuntos (imágenes, PDFs, Excel) enviados
por clientes via WhatsApp u otros canales.

Flujo:
  1. Descarga el archivo desde la API de Meta usando el media_id.
  2. Detecta el tipo (imagen, PDF, Excel, texto).
  3. Extrae el contenido:
     - Imagen  → GPT-4o Vision (base64)
     - PDF     → pdfplumber (texto plano)
     - Excel   → openpyxl (filas como texto)
  4. Devuelve un texto normalizado listo para pasarle a process_message().

Fallback graceful: si falta alguna dependencia o la extracción falla, devuelve
un string de error descriptivo para que el bot pueda responderle al cliente.
"""
from __future__ import annotations

import base64
import io
import logging
import os
from typing import Optional

import requests

logger = logging.getLogger(__name__)

WHATSAPP_GRAPH_BASE = "https://graph.facebook.com/v18.0"

# Prompt enviado al modelo de visión para extraer productos de imágenes
VISION_EXTRACTION_PROMPT = (
    "Analizá esta imagen. Es posible que contenga una lista de productos, "
    "un presupuesto, una orden de compra o una foto de materiales de ferretería. "
    "Extraé todos los productos o ítems que veas, con sus cantidades si están indicadas. "
    "Devolvé solo la lista en texto plano, un ítem por línea, sin explicaciones adicionales. "
    "Si no hay lista de productos, describí brevemente qué ves."
)

PDF_SYSTEM_NOTE = (
    "El cliente envió un PDF. El siguiente es el texto extraído del documento:\n\n"
)

EXCEL_SYSTEM_NOTE = (
    "El cliente envió un archivo Excel. El siguiente es el contenido extraído:\n\n"
)

IMAGE_SYSTEM_NOTE = (
    "El cliente envió una imagen. El siguiente es el contenido extraído:\n\n"
)


class MediaProcessor:
    """Download and extract content from WhatsApp media attachments."""

    def __init__(self, access_token: str, openai_api_key: str = ""):
        self._token      = access_token
        self._openai_key = openai_api_key or os.getenv("OPENAI_API_KEY", "")

    # ─── Public ────────────────────────────────────────────────────────────

    def process_whatsapp_media(self, message: dict) -> Optional[str]:
        """
        Given a raw WhatsApp message dict (from the webhook payload), download
        and extract its content.  Returns a text string ready to feed into the
        bot, or None if the message type is not supported.
        """
        msg_type = message.get("type", "")

        if msg_type == "image":
            media_id  = message.get("image", {}).get("id")
            mime_type = message.get("image", {}).get("mime_type", "image/jpeg")
        elif msg_type == "document":
            media_id  = message.get("document", {}).get("id")
            mime_type = message.get("document", {}).get("mime_type", "application/octet-stream")
        elif msg_type == "audio":
            # Future: transcribir con Whisper
            return None
        else:
            return None

        if not media_id:
            return None

        try:
            content_bytes = self._download_media(media_id)
        except Exception as exc:
            logger.error("media_download_failed media_id=%s: %s", media_id, exc)
            return "No pude descargar el archivo que enviaste. ¿Podés intentarlo de nuevo?"

        return self._extract(content_bytes, mime_type)

    # ─── Extraction ────────────────────────────────────────────────────────

    def _extract(self, data: bytes, mime_type: str) -> str:
        mime = mime_type.lower()

        if mime.startswith("image/"):
            return IMAGE_SYSTEM_NOTE + self._extract_image(data, mime_type)

        if mime == "application/pdf":
            return PDF_SYSTEM_NOTE + self._extract_pdf(data)

        if mime in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "application/excel",
        ) or mime.endswith(".sheet"):
            return EXCEL_SYSTEM_NOTE + self._extract_excel(data)

        if mime.startswith("text/"):
            try:
                return data.decode("utf-8", errors="replace")
            except Exception:
                return "(archivo de texto no legible)"

        logger.warning("media_unsupported_mime_type: %s", mime_type)
        return f"(archivo de tipo {mime_type} — no sé cómo leerlo todavía)"

    def _extract_image(self, data: bytes, mime_type: str) -> str:
        """Use GPT-4o vision to extract product list from an image."""
        if not self._openai_key:
            return "(imagen recibida — activá la API de OpenAI para procesarla automáticamente)"
        try:
            from openai import OpenAI
            client = OpenAI(api_key=self._openai_key)
            b64 = base64.b64encode(data).decode()
            resp = client.chat.completions.create(
                model="gpt-4o",
                max_tokens=1024,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text",      "text": VISION_EXTRACTION_PROMPT},
                        {"type": "image_url", "image_url": {
                            "url":    f"data:{mime_type};base64,{b64}",
                            "detail": "high",
                        }},
                    ],
                }],
            )
            return resp.choices[0].message.content.strip()
        except Exception as exc:
            logger.error("media_vision_extraction_failed: %s", exc)
            return "(no pude leer la imagen automáticamente)"

    def _extract_pdf(self, data: bytes) -> str:
        """Extract text from a PDF using pdfplumber."""
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                pages_text = []
                for page in pdf.pages[:10]:   # máximo 10 páginas
                    text = page.extract_text()
                    if text:
                        pages_text.append(text.strip())
            result = "\n\n".join(pages_text)
            return result if result else "(el PDF no tiene texto extraíble — puede ser una imagen escaneada)"
        except ImportError:
            logger.warning("pdfplumber not installed")
            return "(no se pudo leer el PDF — dependencia faltante)"
        except Exception as exc:
            logger.error("media_pdf_extraction_failed: %s", exc)
            return "(no pude leer el PDF)"

    def _extract_excel(self, data: bytes) -> str:
        """Extract rows from an Excel file using openpyxl."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames[:3]:   # máximo 3 hojas
                ws = wb[sheet_name]
                for row in ws.iter_rows(max_row=200, values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        lines.append(" | ".join(cells))
                if lines:
                    break   # con la primera hoja que tenga datos alcanza
            return "\n".join(lines) if lines else "(el Excel no tiene datos)"
        except ImportError:
            logger.warning("openpyxl not installed")
            return "(no se pudo leer el Excel — dependencia faltante)"
        except Exception as exc:
            logger.error("media_excel_extraction_failed: %s", exc)
            return "(no pude leer el Excel)"

    # ─── Download ──────────────────────────────────────────────────────────

    def _download_media(self, media_id: str) -> bytes:
        """Fetch binary content for a WhatsApp media_id."""
        headers = {"Authorization": f"Bearer {self._token}"}

        # Step 1: resolve the download URL
        info_resp = requests.get(
            f"{WHATSAPP_GRAPH_BASE}/{media_id}",
            headers=headers,
            timeout=15,
        )
        info_resp.raise_for_status()
        download_url = info_resp.json().get("url")
        if not download_url:
            raise ValueError(f"No download URL for media_id={media_id}")

        # Step 2: download the file
        file_resp = requests.get(download_url, headers=headers, timeout=30)
        file_resp.raise_for_status()
        return file_resp.content
